import openpyxl

fileName = 'input/input.xlsx'

def readDevs(firstRow = 4, lastRow = 28):
    devsArray = []
    wb = openpyxl.load_workbook(fileName)
    if "Разработчики" in wb.get_sheet_names():
        ws = wb.get_sheet_by_name("Разработчики")
        colDevId = ws['A']
        colDevType = ws['B']
        colDevName = ws['C']
        colDevHoursPrimary = ws['D']
        colDevHoursSecondary = ws['E']
        colDevHoursExcess = ws['F']
        devsArray = []
        for i in range(firstRow, lastRow):
            devsArray.append([colDevId[i].value, colDevType[i].value, colDevName[i].value, colDevHoursPrimary[i].value, colDevHoursSecondary[i].value, colDevHoursExcess[i].value])
    else:
        print("Нет листа с разработчиками")
    return devsArray


def readTasks(firstRow = 3, lastRow = 338, firstCol = "I", lastCol = "AF"):
    wb = openpyxl.load_workbook(fileName)
    if "57" in wb.get_sheet_names():
        ws = wb.get_sheet_by_name("57")



        for row in range(firstRow, lastRow):




        rowDevId = ws[2]



    return rowDevId


devs = readDevs()
devIds = readTasks()


for dev in devs:
    print(dev)

for cell in devIds:
    print(cell.value)